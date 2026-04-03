#!/usr/bin/env python3
"""
Rebuild Supabase tables from reference Excel (Schema.xlsx).

Creates three tables:
  1. sell_leads_master   (from 'Lead table' sheet)
  2. dcf_leads_master    (from 'DCF table' sheet, deduplicated by lead_id)
  3. dealers_master      (derived from unique dealer codes across both)

Uses the exec_raw_sql RPC endpoint -- same pattern as execute_seed_via_rpc.py.
"""

import json
import sys
import urllib.request
import urllib.error
from datetime import datetime, date

import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
XLSX_PATH = "/Users/a30711/Downloads/Schema.xlsx"
SUPABASE_URL = "https://fdmlyrgiktljuyuthgki.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZkbWx5cmdpa3RsanV5dXRoZ2tpIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzA2OTYyMjQsImV4cCI6MjA4NjI3MjIyNH0."
    "P6td3mqAoKYz6wdgPa9Bs2GytZH4x11n2vuTl6oVb3s"
)
BATCH_SIZE = 50

# ---------------------------------------------------------------------------
# Column mappings  (Excel header -> SQL column name)
# ---------------------------------------------------------------------------

SELL_LEADS_COLUMNS = [
    ("LEAD_ID", "lead_id"),
    ("PRODUCT_APPT_ID", "product_appt_id"),
    ("APPOINTMENT_ID", "appointment_id"),
    ("CX_REG_NO", "cx_reg_no"),
    ("DL_TYPE", "dl_type"),
    ("DL_STATUS", "dl_status"),
    ("MAKE", "make"),
    ("MODEL", "model"),
    ("VARIANT", "variant"),
    ("YEAR", "year"),
    ("DEALER_CODE", "dealer_code"),
    ("DEALER_REGION", "dealer_region"),
    ("GROWTH_DEALER_REGION", "growth_dealer_region"),
    ("VERIFIED", "verified"),
    ("INSPECTION_STORE", "inspection_store"),
    ("INSPECTION_CITY", "inspection_city"),
    ("INSPECTION_REGION", "inspection_region"),
    ("GROWTH_INSPECTING_REGION", "growth_inspecting_region"),
    ("GROWTH_ZONE", "growth_zone"),
    ("LEAD_DATE", "lead_date"),
    ("DEAL_CREATION_DATE", "deal_creation_date"),
    ("ORIGINAL_APPT_DATE", "original_appt_date"),
    ("CURRENT_APPT_DATE", "current_appt_date"),
    ("INSPECTION_DATE", "inspection_date"),
    ("TOKEN_DATE", "token_date"),
    ("STOCKIN_DATE", "stockin_date"),
    ("STOCK_OUT_DATE", "stock_out_date"),
    ("REG_APPT_RANK", "reg_appt_rank"),
    ("REG_INSP_RANK", "reg_insp_rank"),
    ("REG_TOKEN_RANK", "reg_token_rank"),
    ("REG_STOCKIN_RANK", "reg_stockin_rank"),
    ("DL_RANK", "dl_rank"),
    ("FIRST_CANCELLATION_DATE", "first_cancellation_date"),
    ("FIRST_CANCELLATION_FLAG", "first_cancellation_flag"),
    ("CANCELLATION_LOSSES", "cancellation_losses"),
    ("OCB_RUN_COUNT", "ocb_run_count"),
    ("APPOINTMENT_STATUS", "appointment_status"),
    ("LATEST_C24Q", "latest_c24q"),
    ("MAX_C24_QUOTE", "max_c24_quote"),
    ("TARGET_PRICE", "target_price"),
    ("LATEST_OCB_RAISEDAT", "latest_ocb_raisedat"),
    ("VERTICAL", "vertical"),
    ("GS_FLAG", "gs_flag"),
    ("ANS_FLAG", "ans_flag"),
    ("B2B_FLAG", "b2b_flag"),
    ("ELITE_FLAG", "elite_flag"),
    ("SOF_FLAG", "sof_flag"),
    ("FRANCHISE_FLAG", "franchise_flag"),
    ("SELF_BOUGHT_FLAG", "self_bought_flag"),
    ("C2B_ELIGIBLE_INSP", "c2b_eligible_insp"),
    ("LEAD_PURCHASED_DATE", "lead_purchased_date"),
    ("LEAD_PURCHASED", "lead_purchased"),
    ("LEAD_EXPIRED", "lead_expired"),
    ("SELLER_TOKEN", "seller_token"),
    ("INSPECTION_DONE", "inspection_done"),
    ("IM_RAISED", "im_raised"),
    ("PAYMENT_INITIATED", "payment_initiated"),
    ("PAYMENT_OPTION_SELECTED", "payment_option_selected"),
    ("SELF_PAY", "self_pay"),
    ("FULL_ASSIST", "full_assist"),
    ("CREATE_OFFER", "create_offer"),
    ("LEADPURCHASE_PRICE", "leadpurchase_price"),
    ("PAYMENT_DONE", "payment_done"),
    ("REFUNDED", "refunded"),
    ("STOCKINS", "stockins"),
    ("GFD", "gfd"),
    ("INSPECTION_DONE_TIME", "inspection_done_time"),
    ("LEADPRICE", "leadprice"),
    ("SELLERAGREEDPRICE", "selleragreedprice"),
    ("REFUNDED_AT", "refunded_at"),
    ("C2D_STOCKIN_DATE", "c2d_stockin_date"),
    ("SELLER_TOKEN_DATE", "seller_token_date"),
    ("NEGO_LEADS", "nego_leads"),
    ("LEAD_PURCHASE_TIME", "lead_purchase_time"),
    ("SELLER_TOKEN_TIME", "seller_token_time"),
    ("REFUND_TIME", "refund_time"),
    ("BID_AMOUNT", "bid_amount"),
    ("C2D_OCB_FLAG", "c2d_ocb_flag"),
    ("LATEST_C2D_OCB_RAISE_DATE", "latest_c2d_ocb_raise_date"),
    ("C2D_OFFLINE_TOKEN", "c2d_offline_token"),
    ("C2D_OFFLINE_STOCKIN", "c2d_offline_stockin"),
    ("FINAL_C2DTOKEN_DATE", "final_c2dtoken_date"),
    ("FINAL_TOKEN_DATE", "final_token_date"),
    ("FINAL_C2D_SI_DATE", "final_c2d_si_date"),
    ("FINAL_SI_DATE", "final_si_date"),
]

# DCF columns: map Excel header -> SQL column.  Duplicates get unique names.
# We use the Excel column INDEX as the key since headers can repeat.
DCF_COLUMNS_BY_INDEX = {
    0: ("LEAD_ID", "lead_id"),
    1: ("LEAD_CREATION_DATE", "lead_creation_date"),
    2: ("LOGIN_DATE", "login_date"),
    3: ("APPROVAL_DATE", "approval_date"),
    4: ("DISBURSAL_DATETIME", "disbursal_datetime"),
    5: ("CHANNEL", "channel"),
    6: ("DEALER_CITY", "dealer_city"),
    7: ("DEALER_CODEE", "dealer_code"),
    8: ("DEALER_NAME", "dealer_name"),
    9: ("RED_FLAG", "red_flag"),
    10: ("FOS_MAPPING", "fos_mapping"),
    11: ("SM_MAPPING", "sm_mapping"),
    12: ("ASM_MAPPING", "asm_mapping"),
    13: ("RSM_MAPPING", "rsm_mapping"),
    14: ("RM_MAPPING", "rm_mapping"),
    15: ("TL_MAPPING", "tl_mapping"),
    16: ("DS_CHANNEL_", "ds_channel"),
    17: ("NAME", "customer_name"),
    18: ("SYSTEM_LTV", "system_ltv"),
    19: ("FINAL_OFFER_LTV", "final_offer_ltv"),
    20: ("RISK_BUCKET", "risk_bucket"),
    21: ("CIBIL_SCORE_TEMP", "cibil_score"),
    22: ("EMPLOYMENT_DETAILS", "employment_details"),
    23: ("RED_CHANNEL_REASON", "red_channel_reason"),
    24: ("TNC_GENERATED_DATE", "tnc_generated_date"),
    25: ("TNC_ACCEPTED_TIMESTAMP", "tnc_accepted_timestamp"),
    26: ("CASE_STATUS", "case_status"),
    27: ("UNNATI", "unnati"),
    28: ("BANKING_REQUIRED", "banking_required"),
    29: ("Funnel | Loan_State", "funnel_loan_state"),
    30: ("RM_FLOW", "rm_flow"),
    31: ("CC_FOS", "cc_fos"),
    32: ("CC_SM", "cc_sm"),
    33: ("REGISTRATION_NO", "registration_no"),
    34: ("MAKE", "make"),
    35: ("MODEL", "model"),
    36: ("YEAR", "year"),
    37: ("FUEL_TYPE", "fuel_type"),
    38: ("DS_TENURE", "ds_tenure"),
    39: ("DS_ROI", "ds_roi"),
    40: ("SOURCE", "source"),
    41: ("PLATFORM_TYPE", "platform_type"),
    42: ("VALUATION_PRICE", "valuation_price"),
    43: ("BANKING_FLAG", "banking_flag"),
    44: ("ENTRY_POINT", "entry_point"),
    45: ("REGION", "region"),
    46: ("ONBOARD", "onboard"),
    47: ("TL_HEAD", "tl_head"),
    48: ("LOAN_FOR_CHM", "loan_for_chm"),
    49: ("LOAN_FOR_LI", "loan_for_li"),
    50: ("LOAN_FOR_MI", "loan_for_mi"),
    51: ("TOTAL_LOAN_SANCTION", "total_loan_sanction"),
    52: ("LOAN_TENURE", "loan_tenure"),
    53: ("LI_TENURE", "li_tenure"),
    54: ("ROI_PER_ANNUM", "roi_per_annum"),
    55: ("PF_AMOUNT", "pf_amount"),
    56: ("GROSS_DISBURSAL", "gross_disbursal"),
    57: ("HPA_STATUS", "hpa_status"),
    58: ("Case Status", "case_status_2"),
    59: ("TLA", "tla"),
    60: ("Lead_Text", "lead_text"),
    61: ("Lead_date", "lead_date_2"),
    62: ("Disbursal_Revised", "disbursal_revised"),
    63: ("rsa", "rsa"),
    64: ("Buyback", "buyback"),
    65: ("Franchise/ Luxury Franchise/ Normal", "franchise_type"),
    66: ("Channel Revised", "channel_revised"),
    67: ("DS_roi_tla", "ds_roi_tla"),
    68: ("ROI_Loan_Amount", "roi_loan_amount"),
    69: ("Vas Income Own", "vas_income_own"),
    70: ("Vas Income Pmax", "vas_income_pmax"),
    71: ("Mapping Bi City", "mapping_bi_city"),
    72: ("Delhi Flag/DIY", "delhi_flag_diy"),
    73: ("Refinance Dealer", "refinance_dealer"),
    74: ("NBFC Channel", "nbfc_channel"),
    75: ("Disbursal_Timestamp", "disbursal_timestamp"),
    76: ("Revised Region", "revised_region"),
    77: ("Is_lead", "is_lead"),
    78: ("CC FOS", "cc_fos_2"),
    79: ("CC_SM", "cc_sm_2"),
    80: ("Lead Date", "lead_date_clean"),
    81: ("Login Date", "login_date_clean"),
    82: ("Approval Date", "approval_date_clean"),
    83: ("Disbursal Date", "disbursal_date_clean"),
    84: ("Login Date", "login_date_month"),
    85: ("Month", "month_col"),
    86: ("Gross Disbursal Amount", "gross_disbursal_amount"),
    87: ("Rejection Reason normal norms", "rejection_reason"),
    88: ("Final Remarks", "final_remarks"),
    89: ("Current Status", "current_status"),
}

# SQL types per column for sell_leads_master
SELL_LEADS_TYPES = {
    "lead_id": "BIGINT",
    "product_appt_id": "BIGINT",
    "appointment_id": "BIGINT",
    "cx_reg_no": "TEXT",
    "dl_type": "TEXT",
    "dl_status": "TEXT",
    "make": "TEXT",
    "model": "TEXT",
    "variant": "TEXT",
    "year": "INT",
    "dealer_code": "INT",
    "dealer_region": "TEXT",
    "growth_dealer_region": "TEXT",
    "verified": "TEXT",
    "inspection_store": "TEXT",
    "inspection_city": "TEXT",
    "inspection_region": "TEXT",
    "growth_inspecting_region": "TEXT",
    "growth_zone": "TEXT",
    "lead_date": "TIMESTAMPTZ",
    "deal_creation_date": "TIMESTAMPTZ",
    "original_appt_date": "TIMESTAMPTZ",
    "current_appt_date": "TIMESTAMPTZ",
    "inspection_date": "TIMESTAMPTZ",
    "token_date": "TIMESTAMPTZ",
    "stockin_date": "TIMESTAMPTZ",
    "stock_out_date": "TIMESTAMPTZ",
    "reg_appt_rank": "INT",
    "reg_insp_rank": "INT",
    "reg_token_rank": "INT",
    "reg_stockin_rank": "INT",
    "dl_rank": "INT",
    "first_cancellation_date": "TIMESTAMPTZ",
    "first_cancellation_flag": "INT",
    "cancellation_losses": "INT",
    "ocb_run_count": "INT",
    "appointment_status": "TEXT",
    "latest_c24q": "NUMERIC",
    "max_c24_quote": "NUMERIC",
    "target_price": "NUMERIC",
    "latest_ocb_raisedat": "NUMERIC",
    "vertical": "TEXT",
    "gs_flag": "INT",
    "ans_flag": "INT",
    "b2b_flag": "INT",
    "elite_flag": "INT",
    "sof_flag": "INT",
    "franchise_flag": "INT",
    "self_bought_flag": "INT",
    "c2b_eligible_insp": "INT",
    "lead_purchased_date": "TIMESTAMPTZ",
    "lead_purchased": "INT",
    "lead_expired": "INT",
    "seller_token": "INT",
    "inspection_done": "INT",
    "im_raised": "INT",
    "payment_initiated": "INT",
    "payment_option_selected": "INT",
    "self_pay": "INT",
    "full_assist": "INT",
    "create_offer": "INT",
    "leadpurchase_price": "NUMERIC",
    "payment_done": "INT",
    "refunded": "INT",
    "stockins": "INT",
    "gfd": "INT",
    "inspection_done_time": "TIMESTAMPTZ",
    "leadprice": "NUMERIC",
    "selleragreedprice": "NUMERIC",
    "refunded_at": "TIMESTAMPTZ",
    "c2d_stockin_date": "TIMESTAMPTZ",
    "seller_token_date": "TIMESTAMPTZ",
    "nego_leads": "INT",
    "lead_purchase_time": "TIMESTAMPTZ",
    "seller_token_time": "TIMESTAMPTZ",
    "refund_time": "TIMESTAMPTZ",
    "bid_amount": "NUMERIC",
    "c2d_ocb_flag": "INT",
    "latest_c2d_ocb_raise_date": "TIMESTAMPTZ",
    "c2d_offline_token": "INT",
    "c2d_offline_stockin": "INT",
    "final_c2dtoken_date": "TIMESTAMPTZ",
    "final_token_date": "TIMESTAMPTZ",
    "final_c2d_si_date": "TIMESTAMPTZ",
    "final_si_date": "TIMESTAMPTZ",
}

DCF_TYPES = {
    "lead_id": "TEXT",
    "lead_creation_date": "TIMESTAMPTZ",
    "login_date": "TIMESTAMPTZ",
    "approval_date": "TIMESTAMPTZ",
    "disbursal_datetime": "TIMESTAMPTZ",
    "channel": "TEXT",
    "dealer_city": "TEXT",
    "dealer_code": "INT",
    "dealer_name": "TEXT",
    "red_flag": "INT",
    "fos_mapping": "TEXT",
    "sm_mapping": "TEXT",
    "asm_mapping": "TEXT",
    "rsm_mapping": "TEXT",
    "rm_mapping": "TEXT",
    "tl_mapping": "TEXT",
    "ds_channel": "TEXT",
    "customer_name": "TEXT",
    "system_ltv": "NUMERIC",
    "final_offer_ltv": "NUMERIC",
    "risk_bucket": "TEXT",
    "cibil_score": "NUMERIC",
    "employment_details": "TEXT",
    "red_channel_reason": "TEXT",
    "tnc_generated_date": "TIMESTAMPTZ",
    "tnc_accepted_timestamp": "TIMESTAMPTZ",
    "case_status": "TEXT",
    "unnati": "TEXT",
    "banking_required": "INT",
    "funnel_loan_state": "TEXT",
    "rm_flow": "TEXT",
    "cc_fos": "TEXT",
    "cc_sm": "TEXT",
    "registration_no": "TEXT",
    "make": "TEXT",
    "model": "TEXT",
    "year": "INT",
    "fuel_type": "TEXT",
    "ds_tenure": "INT",
    "ds_roi": "NUMERIC",
    "source": "TEXT",
    "platform_type": "TEXT",
    "valuation_price": "NUMERIC",
    "banking_flag": "TEXT",
    "entry_point": "TEXT",
    "region": "TEXT",
    "onboard": "TEXT",
    "tl_head": "TEXT",
    "loan_for_chm": "NUMERIC",
    "loan_for_li": "NUMERIC",
    "loan_for_mi": "NUMERIC",
    "total_loan_sanction": "NUMERIC",
    "loan_tenure": "INT",
    "li_tenure": "INT",
    "roi_per_annum": "NUMERIC",
    "pf_amount": "NUMERIC",
    "gross_disbursal": "NUMERIC",
    "hpa_status": "TEXT",
    "case_status_2": "TEXT",
    "tla": "TEXT",
    "lead_text": "TEXT",
    "lead_date_2": "TIMESTAMPTZ",
    "disbursal_revised": "TEXT",
    "rsa": "INT",
    "buyback": "INT",
    "franchise_type": "TEXT",
    "channel_revised": "TEXT",
    "ds_roi_tla": "TEXT",
    "roi_loan_amount": "NUMERIC",
    "vas_income_own": "NUMERIC",
    "vas_income_pmax": "NUMERIC",
    "mapping_bi_city": "TEXT",
    "delhi_flag_diy": "TEXT",
    "refinance_dealer": "TEXT",
    "nbfc_channel": "TEXT",
    "disbursal_timestamp": "TIMESTAMPTZ",
    "revised_region": "TEXT",
    "is_lead": "INT",
    "cc_fos_2": "TEXT",
    "cc_sm_2": "TEXT",
    "lead_date_clean": "TIMESTAMPTZ",
    "login_date_clean": "TIMESTAMPTZ",
    "approval_date_clean": "TIMESTAMPTZ",
    "disbursal_date_clean": "TIMESTAMPTZ",
    "login_date_month": "TIMESTAMPTZ",
    "month_col": "TIMESTAMPTZ",
    "gross_disbursal_amount": "NUMERIC",
    "rejection_reason": "TEXT",
    "final_remarks": "TEXT",
    "current_status": "TEXT",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def execute_sql(sql_text: str, label: str = "", timeout: int = 300) -> bool:
    """Execute SQL via exec_raw_sql RPC."""
    url = f"{SUPABASE_URL}/rest/v1/rpc/exec_raw_sql"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    body = json.dumps({"sql_text": sql_text}).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.status in (200, 204)
    except urllib.error.HTTPError as e:
        err = e.read().decode("utf-8", errors="replace")
        print(f"  ERROR [{label}] HTTP {e.code}: {err[:500]}")
        return False
    except Exception as e:
        print(f"  ERROR [{label}]: {e}")
        return False


def sql_value(val, col_type: str) -> str:
    """Convert a Python value to a SQL literal string."""
    if val is None:
        return "NULL"

    # Datetime / date objects
    if isinstance(val, (datetime, date)):
        return f"'{val.isoformat()}'"

    # Numeric types stored as TEXT (like cx_reg_no) -- convert to string
    if col_type == "TEXT":
        s = str(val)
        # Handle scientific notation floats that should be text
        if isinstance(val, float):
            # Convert to integer string if it's a whole number (like reg numbers)
            if val == int(val) and abs(val) < 1e20:
                s = str(int(val))
            else:
                s = f"{val:.0f}" if abs(val) < 1e20 else str(val)
        s = s.replace("'", "''")
        return f"'{s}'"

    if col_type == "TIMESTAMPTZ":
        s = str(val)
        if isinstance(val, (datetime, date)):
            s = val.isoformat()
        s = s.replace("'", "''")
        return f"'{s}'"

    if col_type in ("INT", "BIGINT"):
        if isinstance(val, (datetime, date)):
            return "NULL"
        try:
            if isinstance(val, float):
                return str(int(val))
            return str(int(val))
        except (ValueError, TypeError, OverflowError):
            return "NULL"

    if col_type == "NUMERIC":
        if isinstance(val, (datetime, date)):
            return "NULL"
        try:
            # Handle scientific notation strings
            if isinstance(val, str):
                val = float(val)
            return str(val)
        except (ValueError, TypeError):
            return "NULL"

    # Fallback
    s = str(val).replace("'", "''")
    return f"'{s}'"


def build_header_index(ws, mapping_list=None, mapping_dict=None):
    """Build a mapping from Excel column index to (excel_header, sql_col, sql_type).

    For sell_leads we use mapping_list (ordered list of (excel_header, sql_col)).
    For DCF we use mapping_dict keyed by column index.
    """
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    if mapping_dict is not None:
        # DCF: use index-based mapping
        result = {}
        for idx, (excel_hdr, sql_col) in mapping_dict.items():
            if idx < len(headers):
                result[idx] = sql_col
        return result

    # Sell leads: match by header name in order
    result = {}
    header_upper = [h.upper().strip() if h else "" for h in headers]
    for excel_hdr, sql_col in mapping_list:
        target = excel_hdr.upper().strip()
        for i, h in enumerate(header_upper):
            if h == target and i not in result:
                result[i] = sql_col
                break
    return result


# ---------------------------------------------------------------------------
# DDL
# ---------------------------------------------------------------------------

DROP_AND_CREATE_SELL = """
DROP TABLE IF EXISTS sell_leads_master CASCADE;
CREATE TABLE sell_leads_master (
  id SERIAL PRIMARY KEY,
  lead_id BIGINT,
  product_appt_id BIGINT,
  appointment_id BIGINT,
  cx_reg_no TEXT,
  dl_type TEXT,
  dl_status TEXT,
  make TEXT,
  model TEXT,
  variant TEXT,
  year INT,
  dealer_code INT,
  dealer_region TEXT,
  growth_dealer_region TEXT,
  verified TEXT,
  inspection_store TEXT,
  inspection_city TEXT,
  inspection_region TEXT,
  growth_inspecting_region TEXT,
  growth_zone TEXT,
  lead_date TIMESTAMPTZ,
  deal_creation_date TIMESTAMPTZ,
  original_appt_date TIMESTAMPTZ,
  current_appt_date TIMESTAMPTZ,
  inspection_date TIMESTAMPTZ,
  token_date TIMESTAMPTZ,
  stockin_date TIMESTAMPTZ,
  stock_out_date TIMESTAMPTZ,
  reg_appt_rank INT,
  reg_insp_rank INT,
  reg_token_rank INT,
  reg_stockin_rank INT,
  dl_rank INT,
  first_cancellation_date TIMESTAMPTZ,
  first_cancellation_flag INT,
  cancellation_losses INT,
  ocb_run_count INT,
  appointment_status TEXT,
  latest_c24q NUMERIC,
  max_c24_quote NUMERIC,
  target_price NUMERIC,
  latest_ocb_raisedat NUMERIC,
  vertical TEXT,
  gs_flag INT,
  ans_flag INT,
  b2b_flag INT,
  elite_flag INT,
  sof_flag INT,
  franchise_flag INT,
  self_bought_flag INT,
  c2b_eligible_insp INT,
  lead_purchased_date TIMESTAMPTZ,
  lead_purchased INT,
  lead_expired INT,
  seller_token INT,
  inspection_done INT,
  im_raised INT,
  payment_initiated INT,
  payment_option_selected INT,
  self_pay INT,
  full_assist INT,
  create_offer INT,
  leadpurchase_price NUMERIC,
  payment_done INT,
  refunded INT,
  stockins INT,
  gfd INT,
  inspection_done_time TIMESTAMPTZ,
  leadprice NUMERIC,
  selleragreedprice NUMERIC,
  refunded_at TIMESTAMPTZ,
  c2d_stockin_date TIMESTAMPTZ,
  seller_token_date TIMESTAMPTZ,
  nego_leads INT,
  lead_purchase_time TIMESTAMPTZ,
  seller_token_time TIMESTAMPTZ,
  refund_time TIMESTAMPTZ,
  bid_amount NUMERIC,
  c2d_ocb_flag INT,
  latest_c2d_ocb_raise_date TIMESTAMPTZ,
  c2d_offline_token INT,
  c2d_offline_stockin INT,
  final_c2dtoken_date TIMESTAMPTZ,
  final_token_date TIMESTAMPTZ,
  final_c2d_si_date TIMESTAMPTZ,
  final_si_date TIMESTAMPTZ
);
ALTER TABLE sell_leads_master ENABLE ROW LEVEL SECURITY;
"""

DROP_AND_CREATE_DCF = """
DROP TABLE IF EXISTS dcf_leads_master CASCADE;
CREATE TABLE dcf_leads_master (
  id SERIAL PRIMARY KEY,
  lead_id TEXT,
  lead_creation_date TIMESTAMPTZ,
  login_date TIMESTAMPTZ,
  approval_date TIMESTAMPTZ,
  disbursal_datetime TIMESTAMPTZ,
  channel TEXT,
  dealer_city TEXT,
  dealer_code INT,
  dealer_name TEXT,
  red_flag INT,
  fos_mapping TEXT,
  sm_mapping TEXT,
  asm_mapping TEXT,
  rsm_mapping TEXT,
  rm_mapping TEXT,
  tl_mapping TEXT,
  ds_channel TEXT,
  customer_name TEXT,
  system_ltv NUMERIC,
  final_offer_ltv NUMERIC,
  risk_bucket TEXT,
  cibil_score NUMERIC,
  employment_details TEXT,
  red_channel_reason TEXT,
  tnc_generated_date TIMESTAMPTZ,
  tnc_accepted_timestamp TIMESTAMPTZ,
  case_status TEXT,
  unnati TEXT,
  banking_required INT,
  funnel_loan_state TEXT,
  rm_flow TEXT,
  cc_fos TEXT,
  cc_sm TEXT,
  registration_no TEXT,
  make TEXT,
  model TEXT,
  year INT,
  fuel_type TEXT,
  ds_tenure INT,
  ds_roi NUMERIC,
  source TEXT,
  platform_type TEXT,
  valuation_price NUMERIC,
  banking_flag TEXT,
  entry_point TEXT,
  region TEXT,
  onboard TEXT,
  tl_head TEXT,
  loan_for_chm NUMERIC,
  loan_for_li NUMERIC,
  loan_for_mi NUMERIC,
  total_loan_sanction NUMERIC,
  loan_tenure INT,
  li_tenure INT,
  roi_per_annum NUMERIC,
  pf_amount NUMERIC,
  gross_disbursal NUMERIC,
  hpa_status TEXT,
  case_status_2 TEXT,
  tla TEXT,
  lead_text TEXT,
  lead_date_2 TIMESTAMPTZ,
  disbursal_revised TEXT,
  rsa INT,
  buyback INT,
  franchise_type TEXT,
  channel_revised TEXT,
  ds_roi_tla TEXT,
  roi_loan_amount NUMERIC,
  vas_income_own NUMERIC,
  vas_income_pmax NUMERIC,
  mapping_bi_city TEXT,
  delhi_flag_diy TEXT,
  refinance_dealer TEXT,
  nbfc_channel TEXT,
  disbursal_timestamp TIMESTAMPTZ,
  revised_region TEXT,
  is_lead INT,
  cc_fos_2 TEXT,
  cc_sm_2 TEXT,
  lead_date_clean TIMESTAMPTZ,
  login_date_clean TIMESTAMPTZ,
  approval_date_clean TIMESTAMPTZ,
  disbursal_date_clean TIMESTAMPTZ,
  login_date_month TIMESTAMPTZ,
  month_col TIMESTAMPTZ,
  gross_disbursal_amount NUMERIC,
  rejection_reason TEXT,
  final_remarks TEXT,
  current_status TEXT
);
ALTER TABLE dcf_leads_master ENABLE ROW LEVEL SECURITY;
"""

DROP_AND_CREATE_DEALERS = """
DROP TABLE IF EXISTS dealers_master CASCADE;
CREATE TABLE dealers_master (
  id SERIAL PRIMARY KEY,
  dealer_code INT UNIQUE,
  dealer_name TEXT,
  dealer_city TEXT,
  dealer_region TEXT,
  growth_dealer_region TEXT,
  source TEXT
);
ALTER TABLE dealers_master ENABLE ROW LEVEL SECURITY;
"""

RLS_POLICIES = """
DO $$
BEGIN
  IF NOT EXISTS (SELECT 1 FROM pg_policies WHERE policyname = 'anon_read_sell_leads') THEN
    EXECUTE 'CREATE POLICY anon_read_sell_leads ON sell_leads_master FOR SELECT TO anon USING (true)';
  END IF;
  IF NOT EXISTS (SELECT 1 FROM pg_policies WHERE policyname = 'anon_read_dcf_leads') THEN
    EXECUTE 'CREATE POLICY anon_read_dcf_leads ON dcf_leads_master FOR SELECT TO anon USING (true)';
  END IF;
  IF NOT EXISTS (SELECT 1 FROM pg_policies WHERE policyname = 'anon_read_dealers') THEN
    EXECUTE 'CREATE POLICY anon_read_dealers ON dealers_master FOR SELECT TO anon USING (true)';
  END IF;
END $$;
"""


# ---------------------------------------------------------------------------
# Data reading
# ---------------------------------------------------------------------------

def read_sell_leads(wb):
    """Read all rows from the 'Lead table' sheet."""
    ws = wb["Lead table"]
    col_map = build_header_index(ws, mapping_list=SELL_LEADS_COLUMNS)
    print(f"  Mapped {len(col_map)} sell_leads columns")

    # Get ordered SQL column names
    sql_cols = [col_map[i] for i in sorted(col_map.keys())]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = {}
        for idx in col_map:
            cell = row[idx]
            vals[col_map[idx]] = cell.value
        rows.append(vals)

    return sql_cols, rows


def read_dcf_leads(wb):
    """Read DCF rows, deduplicate by lead_id keeping latest lead_creation_date."""
    ws = wb["DCF table"]
    col_map = build_header_index(ws, mapping_dict=DCF_COLUMNS_BY_INDEX)
    print(f"  Mapped {len(col_map)} dcf columns")

    sql_cols = [col_map[i] for i in sorted(col_map.keys())]

    # Read all rows
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = {}
        for idx in col_map:
            if idx < len(row):
                vals[col_map[idx]] = row[idx].value
            else:
                vals[col_map[idx]] = None
        all_rows.append(vals)

    print(f"  Raw DCF rows: {len(all_rows)}")

    # Deduplicate: keep latest lead_creation_date per lead_id
    best = {}  # lead_id -> row
    for r in all_rows:
        lid = r.get("lead_id")
        if lid is None:
            continue
        lcd = r.get("lead_creation_date")
        existing = best.get(lid)
        if existing is None:
            best[lid] = r
        else:
            ex_date = existing.get("lead_creation_date")
            if lcd is not None and (ex_date is None or lcd > ex_date):
                best[lid] = r

    deduped = list(best.values())
    print(f"  After dedup (unique lead_id): {len(deduped)} rows")

    return sql_cols, deduped


# ---------------------------------------------------------------------------
# Batch insert
# ---------------------------------------------------------------------------

def batch_insert(table_name: str, sql_cols: list, rows: list, type_map: dict):
    """Insert rows in batches of BATCH_SIZE."""
    total = len(rows)
    success = 0
    failed = 0

    for batch_start in range(0, total, BATCH_SIZE):
        batch_end = min(batch_start + BATCH_SIZE, total)
        batch = rows[batch_start:batch_end]

        value_rows = []
        for r in batch:
            vals = []
            for col in sql_cols:
                v = r.get(col)
                t = type_map.get(col, "TEXT")
                vals.append(sql_value(v, t))
            value_rows.append(f"({', '.join(vals)})")

        sql = f"INSERT INTO {table_name} ({', '.join(sql_cols)}) VALUES\n"
        sql += ",\n".join(value_rows) + ";"

        label = f"{table_name} rows {batch_start+1}-{batch_end}"
        ok = execute_sql(sql, label=label)
        if ok:
            success += len(batch)
        else:
            failed += len(batch)
            # Try smaller batches on failure
            print(f"  Retrying batch {batch_start+1}-{batch_end} one row at a time...")
            for i, r in enumerate(batch):
                vals = []
                for col in sql_cols:
                    v = r.get(col)
                    t = type_map.get(col, "TEXT")
                    vals.append(sql_value(v, t))
                single_sql = f"INSERT INTO {table_name} ({', '.join(sql_cols)}) VALUES ({', '.join(vals)});"
                single_label = f"{table_name} row {batch_start+i+1}"
                if execute_sql(single_sql, label=single_label):
                    success += 1
                else:
                    failed += 1
                    # Adjust counts: we already counted these as failed
            # Fix double-counting: subtract the batch count we added to failed earlier
            failed -= len(batch)

        pct = (batch_end / total) * 100
        print(f"  [{table_name}] {batch_end}/{total} ({pct:.0f}%) - ok:{success} fail:{failed}", flush=True)

    return success, failed


# ---------------------------------------------------------------------------
# Dealers extraction
# ---------------------------------------------------------------------------

def extract_dealers(sell_rows, dcf_rows):
    """Build dealers_master from unique dealer codes across both datasets."""
    dealers = {}  # dealer_code -> {name, city, region, growth_region, source}

    # From sell_leads
    for r in sell_rows:
        dc = r.get("dealer_code")
        if dc is None:
            continue
        try:
            dc = int(dc)
        except (ValueError, TypeError):
            continue
        if dc not in dealers:
            dealers[dc] = {
                "dealer_name": None,
                "dealer_city": None,
                "dealer_region": r.get("dealer_region"),
                "growth_dealer_region": r.get("growth_dealer_region"),
                "source": "sell_leads",
            }
        else:
            # Fill in missing fields
            if not dealers[dc]["dealer_region"] and r.get("dealer_region"):
                dealers[dc]["dealer_region"] = r.get("dealer_region")
            if not dealers[dc]["growth_dealer_region"] and r.get("growth_dealer_region"):
                dealers[dc]["growth_dealer_region"] = r.get("growth_dealer_region")

    # From DCF
    for r in dcf_rows:
        dc = r.get("dealer_code")
        if dc is None:
            continue
        try:
            dc = int(dc)
        except (ValueError, TypeError):
            continue
        if dc not in dealers:
            dealers[dc] = {
                "dealer_name": r.get("dealer_name"),
                "dealer_city": r.get("dealer_city"),
                "dealer_region": r.get("region"),
                "growth_dealer_region": None,
                "source": "dcf",
            }
        else:
            # Merge
            if dealers[dc]["source"] == "sell_leads":
                dealers[dc]["source"] = "both"
            if not dealers[dc]["dealer_name"] and r.get("dealer_name"):
                dealers[dc]["dealer_name"] = r.get("dealer_name")
            if not dealers[dc]["dealer_city"] and r.get("dealer_city"):
                dealers[dc]["dealer_city"] = r.get("dealer_city")
            if not dealers[dc]["dealer_region"] and r.get("region"):
                dealers[dc]["dealer_region"] = r.get("region")

    return dealers


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("REBUILD FROM REFERENCE - Schema.xlsx -> Supabase")
    print("=" * 60)

    # 1. Load workbook
    print("\n[1/7] Loading workbook...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # 2. Read data
    print("\n[2/7] Reading sell_leads data...")
    sell_cols, sell_rows = read_sell_leads(wb)
    print(f"  Loaded {len(sell_rows)} sell_leads rows, {len(sell_cols)} columns")

    print("\n[3/7] Reading DCF data (with dedup)...")
    dcf_cols, dcf_rows = read_dcf_leads(wb)
    print(f"  Loaded {len(dcf_rows)} deduplicated DCF rows, {len(dcf_cols)} columns")

    wb.close()

    # 3. Drop & create tables
    print("\n[4/7] Creating tables (drop + create)...")
    for label, ddl in [
        ("sell_leads_master", DROP_AND_CREATE_SELL),
        ("dcf_leads_master", DROP_AND_CREATE_DCF),
        ("dealers_master", DROP_AND_CREATE_DEALERS),
    ]:
        print(f"  {label}...", end=" ", flush=True)
        if execute_sql(ddl, label=f"DDL {label}"):
            print("OK")
        else:
            print("FAILED - aborting")
            return 1

    # 4. Insert sell_leads
    print(f"\n[5/7] Inserting sell_leads_master ({len(sell_rows)} rows, batch={BATCH_SIZE})...")
    s_ok, s_fail = batch_insert("sell_leads_master", sell_cols, sell_rows, SELL_LEADS_TYPES)
    print(f"  sell_leads_master: {s_ok} inserted, {s_fail} failed")

    # 5. Insert DCF
    print(f"\n[6/7] Inserting dcf_leads_master ({len(dcf_rows)} rows, batch={BATCH_SIZE})...")
    d_ok, d_fail = batch_insert("dcf_leads_master", dcf_cols, dcf_rows, DCF_TYPES)
    print(f"  dcf_leads_master: {d_ok} inserted, {d_fail} failed")

    # 6. Insert dealers
    print("\n[6b/7] Building and inserting dealers_master...")
    dealers = extract_dealers(sell_rows, dcf_rows)
    print(f"  Found {len(dealers)} unique dealer codes")

    dealer_value_rows = []
    for dc, info in sorted(dealers.items()):
        name = f"'{info['dealer_name'].replace(chr(39), chr(39)+chr(39))}'" if info["dealer_name"] else "NULL"
        city = f"'{info['dealer_city'].replace(chr(39), chr(39)+chr(39))}'" if info["dealer_city"] else "NULL"
        region = f"'{info['dealer_region'].replace(chr(39), chr(39)+chr(39))}'" if info["dealer_region"] else "NULL"
        growth = f"'{info['growth_dealer_region'].replace(chr(39), chr(39)+chr(39))}'" if info["growth_dealer_region"] else "NULL"
        src = f"'{info['source']}'"
        dealer_value_rows.append(f"({dc}, {name}, {city}, {region}, {growth}, {src})")

    # Batch dealers too
    for i in range(0, len(dealer_value_rows), BATCH_SIZE):
        batch = dealer_value_rows[i:i+BATCH_SIZE]
        sql = "INSERT INTO dealers_master (dealer_code, dealer_name, dealer_city, dealer_region, growth_dealer_region, source) VALUES\n"
        sql += ",\n".join(batch) + "\nON CONFLICT (dealer_code) DO NOTHING;"
        execute_sql(sql, label=f"dealers batch {i+1}")

    print(f"  dealers_master: {len(dealers)} rows inserted")

    # 7. RLS policies
    print("\n[7/7] Adding RLS policies...")
    if execute_sql(RLS_POLICIES, label="RLS policies"):
        print("  RLS policies: OK")
    else:
        print("  RLS policies: FAILED (may already exist)")

    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print(f"  sell_leads_master: {s_ok} rows ({s_fail} failed)")
    print(f"  dcf_leads_master:  {d_ok} rows ({d_fail} failed)")
    print(f"  dealers_master:    {len(dealers)} rows")
    print("=" * 60)

    return 0 if (s_fail == 0 and d_fail == 0) else 1


if __name__ == "__main__":
    sys.exit(main())
